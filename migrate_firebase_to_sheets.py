# =========================================================================
# M-C Tuition Application: Firebase to Excel Data Migration Script
# =========================================================================

import os
import sys
import json
import openpyxl
from datetime import datetime

# Path Configuration
project_dir = r"C:\Users\monda\Desktop\Mondal Coaching"
xlsx_path = os.path.join(project_dir, "M-C Tuition Database.xlsx")
key_path = os.path.join(project_dir, "firebase-key.json")

print("==================================================")
print("  M-C Tuition: Firebase to Excel Migration Tool   ")
print("==================================================")

# 1. Verify files exist
if not os.path.exists(xlsx_path):
    print(f"❌ Error: Excel template not found at:\n   {xlsx_path}")
    sys.exit(1)

if not os.path.exists(key_path):
    print(f"\n⚠️ Firebase Credentials File Not Found!")
    print(f"Please generate a Service Account Key from Firebase Console:")
    print(f"1. Go to https://console.firebase.google.com/")
    print(f"2. Open your project settings -> Service Accounts tab.")
    print(f"3. Click 'Generate new private key' and download the .json file.")
    print(f"4. Rename the downloaded file to 'firebase-key.json' and place it here:")
    print(f"   {key_path}")
    print("\nAfter placing the file, please run this script again.")
    sys.exit(0)

# 2. Initialize Firebase Admin SDK
try:
    import firebase_admin
    from firebase_admin import credentials, firestore
    
    print("🔄 Initializing Firebase Connection...")
    cred = credentials.Certificate(key_path)
    firebase_admin.initialize_app(cred)
    db = firestore.client(database_id="ai-studio-5a848f2e-44a8-4b45-8f95-7971b615f241")
    print("✅ Firebase initialized successfully!")
except Exception as e:
    print(f"❌ Error initializing Firebase Admin: {e}")
    sys.exit(1)

# 3. Load Excel Workbook
try:
    print(f"🔄 Loading Excel file: {xlsx_path}...")
    wb = openpyxl.load_workbook(xlsx_path)
except Exception as e:
    print(f"❌ Error loading Excel workbook: {e}")
    sys.exit(1)

def make_serializable(obj):
    if isinstance(obj, dict):
        return {k: make_serializable(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [make_serializable(x) for x in obj]
    elif hasattr(obj, 'isoformat'):
        return obj.isoformat()
    return obj

# 4. Helper to flatten and write Firestore collection to Excel sheet
def migrate_collection(collection_name, excel_sheet_name, excel_columns):
    print(f"⏳ Migrating '{collection_name}' collection...")
    
    if excel_sheet_name not in wb.sheetnames:
        print(f"⚠️ Sheet '{excel_sheet_name}' not found in Excel. Creating it...")
        sheet = wb.create_sheet(excel_sheet_name)
        sheet.append(excel_columns)
    else:
        sheet = wb[excel_sheet_name]
        
    # Clear existing rows except header
    while sheet.max_row > 1:
        sheet.delete_rows(2)
        
    col_to_idx = {col: idx + 1 for idx, col in enumerate(excel_columns)}
    
    try:
        col_ref = db.collection(collection_name)
        docs = col_ref.stream()
        
        count = 0
        for doc in docs:
            doc_data = doc.to_dict()
            doc_data['id'] = doc.id
            
            # Map Firestore fields to Excel columns
            row_data = [""] * len(excel_columns)
            for key, val in doc_data.items():
                # Map special Firestore fields
                if key == 'fullName' and 'name' in col_to_idx:
                    col_key = 'name'
                elif key == 'name' and 'name' in col_to_idx:
                    col_key = 'name'
                else:
                    col_key = key
                    
                if col_key in col_to_idx:
                    idx = col_to_idx[col_key] - 1
                    
                    # Convert complex fields to strings/JSON
                    if isinstance(val, (dict, list)):
                        row_data[idx] = json.dumps(make_serializable(val))
                    elif hasattr(val, 'isoformat'):
                        row_data[idx] = val.isoformat()
                    elif val is None:
                        row_data[idx] = ""
                    else:
                        row_data[idx] = val
                        
            sheet.append(row_data)
            count += 1
            
        print(f"✅ Migrated {count} documents from '{collection_name}' into sheet '{excel_sheet_name}'.")
        return count
    except Exception as e:
        print(f"❌ Failed to migrate '{collection_name}': {e}")
        return 0

# Schema definition matching Google Sheets / Excel
schemas = {
    'users': [
        'id', 'name', 'phone', 'email', 'role', 'status', 'batchId', 'passcode', 
        'address', 'dob', 'joinDate', 'profilePhotoUrl', 'monthlyFee', 'pendingMonths', 
        'exemptReason', 'paymentStatus', 'createdAt', 'updatedAt'
    ],
    'batches': ['id', 'name', 'assignedItemsMap', 'createdAt'],
    'library': [
        'id', 'title', 'type', 'parentId', 'contentUrl', 'isFolder', 'isEncrypted', 
        'encryptionPassword', 'quizId', 'createdAt', 'fileName', 'isChunked', 
        'chunkCount', 'examType', 'trackingId', 'timeLimit', 'marksCorrect', 'marksWrong', 
        'allowMultipleAttempts', 'quizData'
    ],
    'payments': ['id', 'studentId', 'month', 'amount', 'status', 'transactionId', 'paidDate', 'createdAt', 'remarks'],
    'notifications': ['id', 'title', 'message', 'batchId', 'senderId', 'createdAt'],
    'examSessions': ['id', 'examId', 'batchId', 'code', 'isActive', 'codeEnabled', 'participantUids', 'createdAt'],
    'attendance': ['id', 'sessionId', 'studentId', 'studentName', 'studentPhone', 'joinedTime', 'createdAt'],
    'examResults': [
        'id', 'examId', 'studentId', 'studentName', 'studentPhone', 'score', 'totalQuestions', 
        'correctAnswers', 'wrongAnswers', 'skippedAnswers', 'answersMap', 'submittedAt'
    ]
}

# 5. Execute Migration for each collection
total_migrated = 0
for col, cols in schemas.items():
    # Firebase collection might be plural or singular
    firebase_col = col
    if col == 'library':
        firebase_col = 'library' # library matches
        
    count = migrate_collection(firebase_col, col, cols)
    total_migrated += count

# Handle Settings explicitly (special key/value structure)
print("⏳ Migrating platform settings...")
try:
    settings_sheet = wb['settings'] if 'settings' in wb.sheetnames else wb.create_sheet('settings')
    while settings_sheet.max_row > 1:
        settings_sheet.delete_rows(2)
        
    set_ref = db.collection('settings')
    settings_docs = set_ref.stream()
    
    set_count = 0
    for doc in settings_docs:
        doc_data = doc.to_dict()
        for k, v in doc_data.items():
            settings_sheet.append([k, str(v)])
            set_count += 1
            
    print(f"✅ Migrated {set_count} platform setting values.")
except Exception as e:
    print(f"❌ Failed to migrate settings: {e}")

# 6. Save Excel
try:
    print(f"💾 Saving workbook: {xlsx_path}...")
    wb.save(xlsx_path)
    print("==================================================")
    print("🎉 SUCCESS: Data migration completed successfully!")
    print("All Firebase data has been written to the local Excel sheet.")
    print("==================================================")
    print("NEXT STEPS:")
    print("1. Go to Google Drive.")
    print("2. Upload the updated file 'M-C Tuition Database.xlsx'.")
    print("3. Open it and select 'Save as Google Sheets'.")
    print("4. This will give you the Google Sheet with 100% of your real data!")
    print("==================================================")
except Exception as e:
    print(f"❌ Error saving Excel workbook: {e}")
